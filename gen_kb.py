# -*- coding: utf-8 -*-
"""
gen_kb.py — 从《各部门对外承接事项汇总.xlsx》只读提取，生成 bot/data/knowledge.json
用法：python gen_kb.py（在 bot/ 目录下运行，或任意位置运行均可）
注意：只读 Excel，绝不修改源文件。
"""
import json
from pathlib import Path
from collections import Counter

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent          # bot/
EXCEL = BASE.parent / "各部门对外承接事项汇总.xlsx"
OUT = BASE / "data" / "knowledge.json"

SIX_COL_SHEETS = ["香港分部", "内贸销售部", "总务部", "财务部", "自营平台部", "采购部"]
HR_SHEET = "人事部对接人指引"


def main():
    wb = load_workbook(EXCEL, read_only=True)
    entries = []

    # 6 列标准 sheet
    for name in SIX_COL_SHEETS:
        if name not in wb.sheetnames:
            print(f"[warn] 缺少工作表 {name}，跳过")
            continue
        rows = list(wb[name].iter_rows(values_only=True))
        for r in rows[1:]:
            if not r or all((c or "").strip() == "" for c in r):
                continue
            dept = (r[0] or "").strip()
            module = (r[1] or "").strip()
            item = (r[2] or "").strip()
            when = (r[3] or "").strip()
            primary = (r[4] or "").strip()
            backup = (r[5] or "").strip()
            if not item and not primary:
                continue
            entries.append({
                "dept": dept, "module": module, "item": item,
                "when": when, "primary": primary, "backup": backup,
                "source": name,
            })

    # 人事部 3 列 sheet（板块空则沿用上一行）
    if HR_SHEET in wb.sheetnames:
        rows = list(wb[HR_SHEET].iter_rows(values_only=True))
        last_module = ""
        for r in rows[1:]:
            if not r or all((c or "").strip() == "" for c in r):
                continue
            module = (r[0] or "").strip()
            item = (r[1] or "").strip()
            primary = (r[2] or "").strip()
            if not item and not primary:
                continue
            if module:
                last_module = module
            else:
                module = last_module
            entries.append({
                "dept": "人事部", "module": module, "item": item,
                "when": "", "primary": primary, "backup": "",
                "source": HR_SHEET,
            })

    OUT.parent.mkdir(parents=True, exist_ok=True)
    out = {
        "version": f"kb-{len(entries)}-{Path(EXCEL).stat().st_mtime_ns // 10**9}",
        "generated_from": EXCEL.name + "（只读提取）",
        "count": len(entries),
        "entries": entries,
    }
    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"✅ 生成 {len(entries)} 条 → {OUT}")
    for dept, n in Counter(e["dept"] for e in entries).most_common():
        print(f"   {dept}: {n}")


if __name__ == "__main__":
    main()
